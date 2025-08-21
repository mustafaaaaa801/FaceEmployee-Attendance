import face_recognition
import cv2
import pickle
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# تحميل الترميزات
encoding_file = "encodings.pickle"
data = pickle.load(open(encoding_file, "rb"))

# ملف Excel لتسجيل الحضور
attendance_file = "attendance.xlsx"
if not os.path.exists(attendance_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "ID", "Department", "Date", "Time"])
    wb.save(attendance_file)

wb = load_workbook(attendance_file)
ws = wb.active

# تشغيل الكاميرا
cap = cv2.VideoCapture(0)
recognized_employees = set()  # لتجنب تسجيل الحضور أكثر من مرة

print("[INFO] بدء تشغيل الكاميرا...")

while True:
    ret, frame = cap.read()
    if not ret:
        break

    # تحويل إلى RGB
    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

    # تحديد مواقع الوجوه واستخراج الترميزات
    boxes = face_recognition.face_locations(rgb)
    encodings = face_recognition.face_encodings(rgb, boxes)

    names = []
    for encoding in encodings:
        matches = face_recognition.compare_faces(data["encodings"], encoding)
        name = "غير معروف"

        if True in matches:
            matched_idxs = [i for (i, b) in enumerate(matches) if b]
            counts = {}
            for i in matched_idxs:
                name_temp = data["names"][i]
                counts[name_temp] = counts.get(name_temp, 0) + 1
            name = max(counts, key=counts.get)

            # تسجيل الحضور إذا لم يتم تسجيله مسبقًا
            if name not in recognized_employees:
                # استرجاع معلومات الموظف من ملف info.txt
                employee_folder = os.path.join("employees", name)
                info_path = os.path.join(employee_folder, "info.txt")
                emp_id, department = "", ""
                if os.path.exists(info_path):
                    with open(info_path, "r", encoding="utf-8") as f:
                        lines = f.readlines()
                        for line in lines:
                            if line.startswith("ID:"):
                                emp_id = line.replace("ID:", "").strip()
                            elif line.startswith("Department:"):
                                department = line.replace("Department:", "").strip()

                now = datetime.now()
                date_str = now.strftime("%Y-%m-%d")
                time_str = now.strftime("%H:%M:%S")
                ws.append([name, emp_id, department, date_str, time_str])
                wb.save(attendance_file)
                recognized_employees.add(name)
                print(f"تم تسجيل حضور {name} بتاريخ {date_str} الساعة {time_str}")

        names.append(name)

    # رسم المربعات حول الوجوه
    for ((top, right, bottom, left), name) in zip(boxes, names):
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX,
                    0.75, (0, 255, 0), 2)

    cv2.imshow("Face Recognition", frame)

    if cv2.waitKey(1) & 0xFF == ord("q"):
        break

cap.release()
cv2.destroyAllWindows()
